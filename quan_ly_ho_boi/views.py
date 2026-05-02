import json
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpRequest
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.utils.dateparse import parse_date
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Sum
from .models import HoBoi, DatVe, Payment, Review, DichVu, ChiTietDichVu, HinhAnhHoBoi
from .forms import HoBoiForm, DatVeForm, DichVuForm, ChiTietDichVuForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from django.db.models import Avg, Count
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

def gioi_thieu(request):
    return render(request, 'quan_ly_ho_boi/gioi_thieu.html')


# 3. Bản đồ GIS (Trang trống chờ code tool của bạn)
def ban_do(request: HttpRequest):
    return render(request, 'quan_ly_ho_boi/map.html')



# 5. Trang đặt vé cho hồ bơi cụ thể
def dat_ve(request: HttpRequest, ho_boi_id: int):
    ho_boi = get_object_or_404(HoBoi, id=ho_boi_id)
    errors = []
    form_data = {
        'ngay_su_dung': '',
        'gio_su_dung': '',
        'so_luong_nguoi_lon': '1',
        'so_luong_tre_em': '0',
    }
    tong_tien = None
    success_message = None
    dich_vu_list = DichVu.objects.filter(so_luong_kho__gt=0)

    if request.method == 'POST':
        form_data['ngay_su_dung'] = request.POST.get('ngay_su_dung', '')
        form_data['gio_su_dung'] = request.POST.get('gio_su_dung', '') # Lấy giờ bơi
        form_data['so_luong_nguoi_lon'] = request.POST.get('so_luong_nguoi_lon', '1')
        form_data['so_luong_tre_em'] = request.POST.get('so_luong_tre_em', '0')

        # 1. Validate Ngày
        if not form_data['ngay_su_dung']:
            errors.append('Vui lòng chọn ngày sử dụng.')
        try:
            ngay_su_dung = parse_date(form_data['ngay_su_dung'])
            if ngay_su_dung is None: raise ValueError()
        except ValueError:
            errors.append('Ngày sử dụng không hợp lệ.')
            ngay_su_dung = None

        # 2. Validate Giờ & So sánh với giờ mở cửa
        gio_su_dung_obj = None
        if not form_data['gio_su_dung']:
            errors.append('Vui lòng chọn giờ bơi.')
        else:
            try:
                gio_su_dung_obj = datetime.strptime(form_data['gio_su_dung'], '%H:%M').time()
                # Kiểm tra giờ bơi với khung giờ của hồ bơi
                if gio_su_dung_obj < ho_boi.gio_mo_cua or gio_su_dung_obj > ho_boi.gio_dong_cua:
                    errors.append(f'Thất bại: Hồ bơi chỉ mở cửa từ {ho_boi.gio_mo_cua.strftime("%H:%M")} đến {ho_boi.gio_dong_cua.strftime("%H:%M")}.')
            except ValueError:
                errors.append('Giờ sử dụng không hợp lệ.')

        # 3. Validate Số lượng
        try:
            so_luong_nguoi_lon = int(form_data['so_luong_nguoi_lon'])
            so_luong_tre_em = int(form_data['so_luong_tre_em'])
            if so_luong_nguoi_lon < 0 or so_luong_tre_em < 0: raise ValueError()
        except ValueError:
            errors.append('Số lượng không hợp lệ.')

        # Nếu KHÔNG CÓ LỖI thì mới lưu
        if not errors and ngay_su_dung and gio_su_dung_obj:
            current_user = request.user if request.user.is_authenticated else None
            if current_user is None:
                current_user, _ = User.objects.get_or_create(
                    username='guest',
                    defaults={'first_name': 'Khách ẩn danh'}
                )

            dat_ve = DatVe(
                khach_hang=current_user,
                ho_boi=ho_boi,
                ngay_su_dung=ngay_su_dung,
                so_luong_nguoi_lon=so_luong_nguoi_lon,
                so_luong_tre_em=so_luong_tre_em,
            )
            dat_ve.save()
            
            # --- Xử lý dịch vụ đi kèm ---
            for dv in dich_vu_list:
                dv_key = f'dich_vu_{dv.id}'
                if dv_key in request.POST:
                    try:
                        sl_mua = int(request.POST.get(dv_key, 0))
                        if sl_mua > 0 and sl_mua <= dv.so_luong_kho:
                            ChiTietDichVu.objects.create(
                                dat_ve=dat_ve,
                                dich_vu=dv,
                                so_luong=sl_mua,
                                don_gia_thuc_te=dv.don_gia
                            )
                    except ValueError:
                        pass
                        
            return redirect('checkout', datve_id=dat_ve.id)

    else:
        # Tính nhanh tổng tiền hiển thị nếu chỉ có giá vé
        tong_tien = (Decimal(ho_boi.gia_ve_nguoi_lon) * Decimal(form_data['so_luong_nguoi_lon'])
                     + Decimal(ho_boi.gia_ve_tre_em) * Decimal(form_data['so_luong_tre_em']))

    if ho_boi:
        try:
            tong_tien = (Decimal(ho_boi.gia_ve_nguoi_lon) * Decimal(form_data['so_luong_nguoi_lon'])
                         + Decimal(ho_boi.gia_ve_tre_em) * Decimal(form_data['so_luong_tre_em']))
        except Exception:
            tong_tien = None

    return render(request, 'quan_ly_ho_boi/booking.html', {
        'ho_boi': ho_boi,
        'errors': errors,
        'form_data': form_data,
        'tong_tien': tong_tien,
        'success_message': success_message,
        'dich_vu_list': dich_vu_list,
    })


# Sửa lại hàm checkout trong views.py
def checkout(request: HttpRequest, datve_id: int):
    dat_ve = get_object_or_404(DatVe, id=datve_id)
    payment_message = None
    payment_method = 'Tiền mặt'

    if request.method == 'POST':
        method = request.POST.get('payment_method', 'cash')
        labels = {
            'cash': 'Tiền mặt',
            'card': 'Thẻ ngân hàng',
            'mobile': 'Thanh toán di động',
        }
        payment_method = labels.get(method, 'Tiền mặt')
        
        # LOGIC MỚI: Tùy chỉnh trạng thái dựa trên phương thức
        trang_thai_thanh_toan = 'Đang chờ' if method == 'cash' else 'Hoàn thành'

        payment, created = Payment.objects.get_or_create(
            dat_ve=dat_ve,
            defaults={
                'phuong_thuc': payment_method,
                'so_tien': dat_ve.tong_thanh_toan_cuoi, # Áp dụng tổng cuối cùng bao gồm dịch vụ
                'trang_thai': trang_thai_thanh_toan, # Áp dụng trạng thái mới
            }
        )
        if not created:
            payment.phuong_thuc = payment_method
            payment.so_tien = dat_ve.tong_thanh_toan_cuoi
            payment.trang_thai = trang_thai_thanh_toan
            payment.save()

        # Hiển thị câu thông báo khác nhau cho khách hàng
        if trang_thai_thanh_toan == 'Đang chờ':
            payment_message = f"Ghi nhận đặt vé #{dat_ve.id}. Vui lòng thanh toán Tiền mặt tại quầy để nhân viên xác nhận vào cổng!"
        else:
            payment_message = f"Thanh toán thành công bằng {payment_method}. Vé đặt mã #{dat_ve.id} đã được xác nhận."

    return render(request, 'quan_ly_ho_boi/checkout.html', {
        'dat_ve': dat_ve,
        'payment_message': payment_message,
        'payment_method': payment_method,
    })

# 6. Lịch sử thanh toán

def lich_su_thanh_toan(request: HttpRequest):
    lich_su = Payment.objects.select_related('dat_ve__ho_boi', 'dat_ve__khach_hang').order_by('-ngay_thanh_toan')
    return render(request, 'quan_ly_ho_boi/payment_history.html', {
        'lich_su': lich_su,
    })


# 8. Trang admin tùy chỉnh (Tổng hợp quản lý)
@login_required
def admin_home(request: HttpRequest):
    # Kiểm tra quyền truy cập Admin
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
    
    # --- 1. DỮ LIỆU THỐNG KÊ TỔNG QUAN (DASHBOARD) ---
    tong_ho_boi = HoBoi.objects.count()
    tong_ve_dat = DatVe.objects.count()
    ho_dang_mo = HoBoi.objects.filter(trang_thai='MO').count()
    tong_doanh_thu = Payment.objects.filter(trang_thai='Hoàn thành').aggregate(total=Sum('so_tien'))['total'] or 0
    
    # Lấy 5 hồ bơi mới nhất hiện nhanh ở Dashboard
    ho_boi_gan_day = HoBoi.objects.order_by('-id')[:5]
    
    # Danh sách chờ thu tiền mặt (thường để ít nên không cần phân trang)
    thanh_toan_cho = Payment.objects.filter(trang_thai='Đang chờ').select_related('dat_ve__khach_hang', 'dat_ve__ho_boi')

    # --- 2. PHÂN TRANG DANH SÁCH HỒ BƠI (TAB QUẢN LÝ HỒ BƠI) ---
    danh_sach_ho_full = HoBoi.objects.all().order_by('-id')
    paginator_ho = Paginator(danh_sach_ho_full, 10) # 10 hồ bơi mỗi trang
    page_ho_number = request.GET.get('page_ho')
    danh_sach_ho = paginator_ho.get_page(page_ho_number)

    # --- 3. PHÂN TRANG DỊCH VỤ (TAB QUẢN LÝ DỊCH VỤ) ---
    danh_sach_dich_vu_full = DichVu.objects.all().order_by('-id')
    paginator_dich_vu = Paginator(danh_sach_dich_vu_full, 10)
    page_dich_vu_number = request.GET.get('page_dv')
    danh_sach_dich_vu = paginator_dich_vu.get_page(page_dich_vu_number)

    # --- 3. PHÂN TRANG TẤT CẢ VÉ ĐẶT (TAB QUẢN LÝ VÉ ĐẶT) ---
    tat_ca_ve_dat_full = DatVe.objects.select_related('ho_boi', 'khach_hang').order_by('-ngay_dat')
    paginator_ve = Paginator(tat_ca_ve_dat_full, 7) # 10 vé mỗi trang
    page_ve_number = request.GET.get('page_ve')
    ve_dat_gan_day = paginator_ve.get_page(page_ve_number)

    danh_sach_danh_gia_full = Review.objects.select_related('ho_boi', 'user').order_by('-created_at')
    paginator_danh_gia = Paginator(danh_sach_danh_gia_full, 10) # 10 bình luận/trang
    page_danh_gia_number = request.GET.get('page_danh_gia')
    danh_sach_danh_gia = paginator_danh_gia.get_page(page_danh_gia_number)

    # --- 4. ĐÓNG GÓI CONTEXT ---
    context = {
        # Thống kê Dashboard
        'tong_ho_boi': tong_ho_boi,
        'tong_ve_dat': tong_ve_dat,
        'ho_dang_mo': ho_dang_mo,
        'tong_doanh_thu': tong_doanh_thu,
        'ho_boi_gan_day': ho_boi_gan_day,
        'thanh_toan_cho': thanh_toan_cho,
        
        # Dữ liệu phân trang (Object Page)
        'danh_sach_ho': danh_sach_ho,
        've_dat_gan_day': ve_dat_gan_day,
        
        # Biến đếm tổng để hiện trên Badge (Huy hiệu)
        'tong_ho_count': paginator_ho.count,
        'tong_ve_count': paginator_ve.count,

        'danh_sach_danh_gia': danh_sach_danh_gia,
        'tong_danh_gia_count': paginator_danh_gia.count,
        
        'danh_sach_dich_vu': danh_sach_dich_vu,
        'tong_dich_vu_count': paginator_dich_vu.count,
    }
    
    return render(request, 'quan_ly_ho_boi/admin_panel.html', context)

# 8.1. Login cho admin panel
def login_admin(request: HttpRequest):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('admin_panel')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('admin_panel')
        else:
            messages.error(request, 'Tên đăng nhập hoặc mật khẩu không đúng, hoặc bạn không có quyền truy cập.')
    
    return render(request, 'quan_ly_ho_boi/login_admin.html')

# 8.2. Logout cho admin panel
def logout_admin(request: HttpRequest):
    logout(request)
    return redirect('login_admin')

# 9. Chức năng Xóa hồ bơi
def xoa_ho_boi(request: HttpRequest, ho_boi_id : int):
    ho_boi = get_object_or_404(HoBoi, id=ho_boi_id)
    if request.method == 'POST':
        ho_boi.delete()
        return redirect('admin_panel')
    return redirect('admin_panel')

@login_required
def xoa_dat_ve(request: HttpRequest, datve_id: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)

    dat_ve = get_object_or_404(DatVe, id=datve_id)
    if request.method == 'POST':
        dat_ve.delete()
        messages.success(request, 'Vé đặt đã được xóa.')
        return redirect('admin_panel')
    return redirect('admin_panel')

# 10. Tạo hồ bơi mới trong admin
@login_required
def tao_ho_boi(request: HttpRequest):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)

    if request.method == 'POST':
        form = HoBoiForm(request.POST, request.FILES)
        if form.is_valid():
            ho_boi = form.save()
            
            # Lưu ảnh bộ sưu tập
            for f in request.FILES.getlist('danh_sach_anh_phu'):
                HinhAnhHoBoi.objects.create(ho_boi=ho_boi, hinh_anh=f)
                
            messages.success(request, 'Hồ bơi đã được tạo thành công.')
            return redirect('admin_panel')
    else:
        form = HoBoiForm()

    return render(request, 'quan_ly_ho_boi/admin_pool_form.html', {
        'form': form,
    })

# 11. Chỉnh sửa hồ bơi trong admin
@login_required
def chinh_sua_ho_boi(request: HttpRequest, ho_boi_id: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)

    ho_boi = get_object_or_404(HoBoi, id=ho_boi_id)
    if request.method == 'POST':
        form = HoBoiForm(request.POST, request.FILES, instance=ho_boi)
        if form.is_valid():
            ho_boi = form.save()
            
            # Lưu ảnh bộ sưu tập mới thêm
            for f in request.FILES.getlist('danh_sach_anh_phu'):
                HinhAnhHoBoi.objects.create(ho_boi=ho_boi, hinh_anh=f)
                
            messages.success(request, 'Thông tin hồ bơi đã được cập nhật.')
            return redirect('admin_panel')
    else:
        form = HoBoiForm(instance=ho_boi)

    return render(request, 'quan_ly_ho_boi/admin_pool_form.html', {
        'form': form,
    })

@login_required
def chinh_sua_dat_ve(request: HttpRequest, datve_id: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)

    dat_ve = get_object_or_404(DatVe, id=datve_id)
    if request.method == 'POST':
        form = DatVeForm(request.POST, instance=dat_ve)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thông tin vé đã được cập nhật.')
            return redirect('admin_panel')
    else:
        form = DatVeForm(instance=dat_ve)

    return render(request, 'quan_ly_ho_boi/admin_edit_booking.html', {
        'form': form,
        'dat_ve': dat_ve,
        'tat_ca_dich_vu': DichVu.objects.all(),
    })

# 12. Trang chủ công khai (Home Page)
def home_page(request: HttpRequest):
    """Trang chủ cho khách hàng công khai"""
    danh_sach_ho = HoBoi.objects.all()
    ho_dang_mo = HoBoi.objects.filter(trang_thai='MO').count()
    context = {
        'danh_sach_ho': danh_sach_ho,
        'ho_dang_mo': ho_dang_mo,
    }
    return render(request, 'quan_ly_ho_boi/home.html', context)

# 11. Trang cá nhân (User Profile)
@login_required(login_url='/login/')
def user_profile(request: HttpRequest):
    """Trang thông tin cá nhân của người dùng (Bắt buộc đăng nhập)"""
    user = request.user
    # Lịch sử vé đặt của người dùng
    lich_su_ve = DatVe.objects.filter(khach_hang=user).select_related('ho_boi').order_by('-ngay_dat')[:10]
    # Lịch sử thanh toán của người dùng
    lich_su_thanh_toan = Payment.objects.filter(dat_ve__khach_hang=user).select_related('dat_ve__ho_boi').order_by('-ngay_thanh_toan')[:5]
    
    context = {
        'user': user,
        'lich_su_ve': lich_su_ve,
        'lich_su_thanh_toan': lich_su_thanh_toan,
        'tong_ve_da_dat': lich_su_ve.count(),
    }
    return render(request, 'quan_ly_ho_boi/profile.html', context)

# Custom 404 page handler
def page_not_found_view(request, exception=None):
    return render(request, 'quan_ly_ho_boi/404.html', status=404)

# 12. Đăng nhập cho user bình thường
def login_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('profile')
        else:
            return render(request, 'quan_ly_ho_boi/login_user.html', {'error': 'Tên đăng nhập hoặc mật khẩu không đúng.'})
    return render(request, 'quan_ly_ho_boi/login_user.html')

# 13. Đăng xuất user
def logout_user(request):
    logout(request)
    return redirect('home')

# 14. Đăng ký user
def register_user(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = False
            user.is_superuser = False
            user.save()
            return redirect('login_user')
    else:
        form = UserCreationForm()
    return render(request, 'quan_ly_ho_boi/register.html', {'form': form})

# 6. Lưu xử lí dữ liệu
@csrf_exempt
def luu_ho_boi(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            # Log kiểm tra dữ liệu đầu vào
            print("Data nhận được:", data)
            
            ho_boi_moi = HoBoi.objects.create(
                ten_ho=data.get('name'),
                vi_do=float(data.get('lat')),
                kinh_do=float(data.get('lng')),
                # Cung cấp giá trị mặc định cho các trường bắt buộc trong Model
                dia_chi="Chưa xác định", 
                do_sau=1.5,
                suc_chua=50,
                trang_thai='MO'
            )
            return JsonResponse({'status': 'success', 'id': ho_boi_moi.id})
        except Exception as e:
            print("Lỗi Server:", str(e)) # Xem lỗi cụ thể tại Terminal
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'failed'}, status=400)
# 7. Xuất dữ liệu hồ
def get_pools(request):
    pools = HoBoi.objects.all()
    data = []
    for p in pools:
        data.append({
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [p.kinh_do, p.vi_do]
            },
            "properties": {
                "id": p.id,
                "name": p.ten_ho,
                "address": p.dia_chi,
                "status_code": p.trang_thai,
                "status": p.get_trang_thai_display(),
                "depth": p.do_sau,
                "capacity": p.suc_chua,
                "price_adult": str(p.gia_ve_nguoi_lon),
                "price_child": str(p.gia_ve_tre_em),
                "image": p.hinh_anh.url if p.hinh_anh else '',
                'gallery': [anh.hinh_anh.url for anh in p.danh_sach_hinh_anh.all()] if hasattr(HoBoi, 'danh_sach_hinh_anh') else []
            }
        })
    return JsonResponse({"type": "FeatureCollection", "features": data})

def pool_list(request):
    # Lấy tất cả HoBoi, tính điểm trung bình (avg_rating) 
    # và số lượng đánh giá (review_count) cho mỗi hồ
    pools = HoBoi.objects.annotate(
        avg_rating=Avg('reviews__rating'),
        # Thêm distinct=True vào đây để đảm bảo đếm chính xác, không bị nhân bản dòng
        review_count=Count('reviews', distinct=True) 
    ).order_by('-avg_rating').distinct().all()
    
    return render(request, 'quan_ly_ho_boi/pool_list.html', {'pools': pools})

@login_required
def add_review(request, ho_boi_id):
    ho_boi = get_object_or_404(HoBoi, id=ho_boi_id)
    
    if request.method == 'POST':
        # KIỂM TRA: Nếu user đã đánh giá hồ bơi này rồi thì báo lỗi và chặn lại
        if Review.objects.filter(ho_boi=ho_boi, user=request.user).exists():
            messages.error(request, "Bạn đã đánh giá hồ bơi này rồi. Mỗi tài khoản chỉ được đánh giá 1 lần!")
            return redirect('pool_list')

        rating = request.POST.get('rating')
        comment = request.POST.get('comment')
        
        Review.objects.create(
            ho_boi=ho_boi,
            user=request.user,
            rating=rating,
            comment=comment
        )
        messages.success(request, "Cảm ơn bạn đã gửi đánh giá!")
        
    return redirect('pool_list')
def submit_review(request):
    if request.method == 'POST':
        if not request.user.is_authenticated:
            return JsonResponse({'status': 'error', 'message': 'Bạn cần đăng nhập!'}, status=403)
            
        try:
            data = json.loads(request.body)
            # Tìm hồ bơi theo ID
            ho_boi = HoBoi.objects.get(id=data['pool_id'])
            
            # Tạo đánh giá liên kết với HoBoi
            Review.objects.create(
                ho_boi=ho_boi,
                user=request.user,
                rating=data['rating'],
                comment=data['comment']
            )
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
            
    return JsonResponse({'status': 'error', 'message': 'Yêu cầu không hợp lệ'}, status=400)

def test_payment_api(request):
    """API giả lập cổng thanh toán ngân hàng"""
    if request.method == 'POST':
        try:
            # Lấy dữ liệu từ Frontend gửi lên
            data = json.loads(request.body)
            so_the = data.get('card_number')
            
            # GIẢ LẬP LOGIC NGÂN HÀNG:
            # Quy ước: Chỉ thẻ có đuôi "9999" mới thanh toán thành công, còn lại báo lỗi (để test cả 2 trường hợp)
            if so_the and so_the.endswith('9999'):
                return JsonResponse({
                    'status': 'success', 
                    'transaction_id': 'TEST_' + str(request.user.id) + '_8888',
                    'message': 'Thanh toán qua thẻ thành công!'
                })
            else:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Thẻ không hợp lệ hoặc số dư không đủ. (Gợi ý: Dùng thẻ đuôi 9999)'
                })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
            
    return JsonResponse({'status': 'error', 'message': 'Yêu cầu không hợp lệ'}, status=400)

@login_required
def xac_nhan_thanh_toan(request: HttpRequest, payment_id: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
        
    payment = get_object_or_404(Payment, id=payment_id)
    if request.method == 'POST':
        payment.trang_thai = 'Hoàn thành'
        payment.save()
        messages.success(request, f'Đã thu tiền và xác nhận thành công cho vé #{payment.dat_ve.id}!')
        
    return redirect('admin_panel')

# Thêm vào cuối file views.py
@login_required
def in_hoa_don(request: HttpRequest, datve_id: int):
    # Lấy vé đặt
    dat_ve = get_object_or_404(DatVe, id=datve_id)
    
    # Bảo mật: Chỉ Admin hoặc chính người mua vé đó mới được xem/in hóa đơn
    if not request.user.is_staff and request.user != dat_ve.khach_hang:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
        
    return render(request, 'quan_ly_ho_boi/invoice.html', {
        'dat_ve': dat_ve,
    })

@login_required
def xoa_danh_gia(request: HttpRequest, review_id: int):
    # Kiểm tra quyền Admin
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)

    review = get_object_or_404(Review, id=review_id)
    if request.method == 'POST':
        review.delete()
        messages.success(request, 'Đã xóa bình luận thành công.')
        return redirect('admin_panel')
    return redirect('admin_panel')

@login_required
def xuat_excel_ve_dat(request: HttpRequest):
    # Kiểm tra quyền Admin
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)

    # Khởi tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh Sách Vé Đặt"

    # 1. Tạo Dòng Tiêu Đề
    columns = ['Mã vé', 'Khách hàng', 'Hồ bơi', 'Ngày đặt', 'Ngày sử dụng', 'Trạng thái TT', 'Thành tiền (VNĐ)']
    ws.append(columns)

    # Định dạng in đậm và căn giữa cho Tiêu đề
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    # Tùy chỉnh độ rộng cột cho đẹp
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20

    # 2. Lấy dữ liệu từ Database
    ve_dat_list = DatVe.objects.select_related('ho_boi', 'khach_hang').order_by('-ngay_dat')
    
    tong_doanh_thu_thuc_thu = 0

    for ve in ve_dat_list:
        # Xử lý trạng thái an toàn (tránh lỗi nếu vé chưa có payment)
        if hasattr(ve, 'payment'):
            trang_thai = ve.payment.trang_thai
        else:
            trang_thai = 'Chưa Thanh Toán'

        # Chỉ cộng vào tổng doanh thu nếu vé đã 'Hoàn thành'
        if trang_thai == 'Hoàn thành':
            tong_doanh_thu_thuc_thu += float(ve.tong_thanh_toan_cuoi)

        # Thêm từng dòng dữ liệu
        ws.append([
            f"#{ve.id}",
            ve.khach_hang.username,
            ve.ho_boi.ten_ho,
            ve.ngay_dat.strftime("%d/%m/%Y %H:%M"),
            ve.ngay_su_dung.strftime("%d/%m/%Y"),
            trang_thai,
            float(ve.tong_thanh_toan_cuoi)
        ])

    # 3. Thêm dòng Tổng Doanh Thu ở cuối
    ws.append([]) # Thêm 1 dòng trống cho dễ nhìn
    
    # Dòng tính tổng
    ws.append(["", "", "", "", "", "TỔNG THỰC THU:", tong_doanh_thu_thuc_thu])
    last_row = ws.max_row
    
    # Định dạng in đậm và tô màu đỏ chữ cho dòng Tổng
    ws.cell(row=last_row, column=6).font = Font(bold=True, color="FF0000")
    ws.cell(row=last_row, column=7).font = Font(bold=True, color="FF0000")

    # 4. Trả file về cho trình duyệt tải xuống
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Danh_Sach_Ve_Dat.xlsx"'
    wb.save(response)
    return response

# ==========================================
# MODULE QUẢN LÝ DỊCH VỤ VÀ DỤNG CỤ
# ==========================================

@login_required
def dichvu_list(request: HttpRequest):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
    dich_vu_list = DichVu.objects.all()
    return render(request, 'quan_ly_ho_boi/admin_dichvu_list.html', {'dich_vu_list': dich_vu_list})

@login_required
def dichvu_create(request: HttpRequest):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
    if request.method == 'POST':
        form = DichVuForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thêm dịch vụ thành công.')
            return redirect('/admin-panel/?tab=tab-services')
    else:
        form = DichVuForm()
    return render(request, 'quan_ly_ho_boi/admin_dichvu_form.html', {'form': form, 'title': 'Thêm Dịch Vụ Mới'})

@login_required
def dichvu_update(request: HttpRequest, pk: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
    dich_vu = get_object_or_404(DichVu, pk=pk)
    if request.method == 'POST':
        form = DichVuForm(request.POST, instance=dich_vu)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật dịch vụ thành công.')
            return redirect('/admin-panel/?tab=tab-services')
    else:
        form = DichVuForm(instance=dich_vu)
    return render(request, 'quan_ly_ho_boi/admin_dichvu_form.html', {'form': form, 'title': 'Cập Nhật Dịch Vụ'})

@login_required
def dichvu_delete(request: HttpRequest, pk: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
    dich_vu = get_object_or_404(DichVu, pk=pk)
    if request.method == 'POST':
        dich_vu.delete()
        messages.success(request, 'Xóa dịch vụ thành công.')
        return redirect('/admin-panel/?tab=tab-services')
    return redirect('/admin-panel/?tab=tab-services')

@login_required
def add_chitiet_dichvu(request: HttpRequest, datve_id: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
    dat_ve = get_object_or_404(DatVe, pk=datve_id)
    
    if request.method == 'POST':
        dich_vu_id = request.POST.get('dich_vu')
        so_luong = int(request.POST.get('so_luong', 1))
        
        if dich_vu_id:
            dich_vu = get_object_or_404(DichVu, pk=dich_vu_id)
            if dich_vu.so_luong_kho < so_luong:
                messages.error(request, f'Trong kho chỉ còn {dich_vu.so_luong_kho} {dich_vu.ten_dich_vu}.')
            else:
                ChiTietDichVu.objects.create(
                    dat_ve=dat_ve,
                    dich_vu=dich_vu,
                    so_luong=so_luong
                )
                messages.success(request, 'Đã thêm dịch vụ vào hóa đơn.')
        else:
            messages.error(request, 'Vui lòng chọn một dịch vụ.')
            
    return redirect('chinh_sua_dat_ve', datve_id=dat_ve.id)

@login_required
def update_chitiet_dichvu_status(request: HttpRequest, chitiet_id: int):
    if not request.user.is_staff:
        return render(request, 'quan_ly_ho_boi/403.html', status=403)
    chitiet = get_object_or_404(ChiTietDichVu, pk=chitiet_id)
    if request.method == 'POST':
        new_status = request.POST.get('trang_thai')
        if new_status in dict(ChiTietDichVu.TRANG_THAI_CHOICES):
            chitiet.trang_thai = new_status
            chitiet.save()
            messages.success(request, 'Cập nhật trạng thái thành công.')
        else:
            messages.error(request, 'Trạng thái không hợp lệ.')
    return redirect('chinh_sua_dat_ve', datve_id=chitiet.dat_ve.id)